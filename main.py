from __future__ import annotations

import hashlib
import importlib.util
from collections import defaultdict
from io import BytesIO
from zipfile import BadZipFile, ZipFile

import pandas as pd
import streamlit as st

from processor import build_consolidated_dataframe, build_policy_comparison_dataframe, process_expense_file

st.set_page_config(page_title="Expense Consolidation Tool", layout="wide")
st.title("Dynamic Employee Expense Consolidation")
st.caption(
    "Upload employee sheets (.xlsx/.xls) or ZIP bundles. "
    "The app auto-detects dynamic expense heads and consolidates employee-wise."
)


def _unique_name(name: str, seen: dict[str, int]) -> str:
    seen[name] += 1
    if seen[name] == 1:
        return name

    if "." in name:
        stem, ext = name.rsplit(".", 1)
        return f"{stem}__{seen[name]}.{ext}"
    return f"{name}__{seen[name]}"


def _expand_uploads(uploaded_files) -> tuple[list[tuple[str, bytes]], dict[str, str], list[str]]:
    expanded_files: list[tuple[str, bytes]] = []
    precheck_errors: dict[str, str] = {}
    notes: list[str] = []

    name_counts: dict[str, int] = defaultdict(int)
    content_hashes: set[str] = set()

    for uploaded in uploaded_files:
        uploaded.seek(0)
        blob = uploaded.read()

        if not blob:
            precheck_errors[uploaded.name] = "Uploaded file is empty."
            continue

        # Optional duplicate-content protection to avoid repeated uploads.
        digest = hashlib.sha256(blob).hexdigest()
        if digest in content_hashes:
            precheck_errors[uploaded.name] = "Duplicate file content detected; skipped."
            continue
        content_hashes.add(digest)

        if uploaded.name.lower().endswith(".zip"):
            try:
                with ZipFile(BytesIO(blob)) as archive:
                    members = [
                        m
                        for m in archive.infolist()
                        if not m.is_dir() and m.filename.lower().endswith((".xlsx", ".xls"))
                    ]
                    if not members:
                        precheck_errors[uploaded.name] = "ZIP contains no .xlsx/.xls files."
                        continue

                    for member in members:
                        member_bytes = archive.read(member)
                        if not member_bytes:
                            precheck_errors[member.filename] = "ZIP entry is empty."
                            continue

                        member_hash = hashlib.sha256(member_bytes).hexdigest()
                        if member_hash in content_hashes:
                            precheck_errors[member.filename] = "Duplicate file content detected in ZIP; skipped."
                            continue
                        content_hashes.add(member_hash)

                        base_name = member.filename.split("/")[-1] or member.filename
                        unique_name = _unique_name(base_name, name_counts)
                        if unique_name != base_name:
                            notes.append(f"Duplicate filename renamed: {base_name} -> {unique_name}")
                        expanded_files.append((unique_name, member_bytes))
            except BadZipFile:
                precheck_errors[uploaded.name] = "Invalid or corrupt ZIP archive."
            continue

        unique_name = _unique_name(uploaded.name, name_counts)
        if unique_name != uploaded.name:
            notes.append(f"Duplicate filename renamed: {uploaded.name} -> {unique_name}")
        expanded_files.append((unique_name, blob))

    return expanded_files, precheck_errors, notes


@st.cache_data(show_spinner=False)
def _process_file_cached(file_name: str, file_bytes: bytes):
    return process_expense_file(file_name, file_bytes)


def _pick_excel_engine() -> str | None:
    """Select an available Excel writer engine without importing optional deps at module import time."""
    if importlib.util.find_spec("xlsxwriter") is not None:
        return "xlsxwriter"
    if importlib.util.find_spec("openpyxl") is not None:
        return "openpyxl"
    return None


def _style_output_workbook(writer, engine: str):
    if engine == "xlsxwriter":
        workbook = writer.book
        header_format = workbook.add_format(
            {
                "font_name": "Century Gothic",
                "font_size": 10,
                "bold": True,
                "font_color": "#FFFFFF",
                "bg_color": "#1F4E78",
            }
        )
        body_format = workbook.add_format({"font_name": "Century Gothic", "font_size": 10})

        for sheet_name, df in writer.sheets.items():
            worksheet = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(df.columns):
                values = [str(col_name)] + ["" if v is None else str(v) for v in df[col_name]]
                width = min(max(max(len(v) for v in values) + 2, 10), 60)
                worksheet.set_column(col_idx, col_idx, width, body_format)
                worksheet.write(0, col_idx, str(col_name), header_format)

    elif engine == "openpyxl":
        from openpyxl.styles import Font, PatternFill

        workbook = writer.book
        base_font = Font(name="Century Gothic", size=10)
        header_font = Font(name="Century Gothic", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = base_font

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            for column in sheet.columns:
                max_len = 0
                col_letter = column[0].column_letter
                for cell in column:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

uploaded_files = st.file_uploader(
    "Upload Employee Expense Sheets",
    type=["xlsx", "xls", "zip"],
    accept_multiple_files=True,
)

if uploaded_files and st.button("Generate Consolidated Report", type="primary"):
    files_to_process, precheck_errors, notes = _expand_uploads(uploaded_files)

    if notes:
        st.warning("\n".join(notes))

    if precheck_errors:
        st.error("Upload validation issues found:")
        st.json(precheck_errors)

    if not files_to_process:
        st.stop()

    progress = st.progress(0.0)

    successful_rows: list[dict] = []
    expense_heads: set[str] = set()
    failed_files: dict[str, str] = dict(precheck_errors)

    for index, (file_name, file_bytes) in enumerate(files_to_process, start=1):
        result = _process_file_cached(file_name, file_bytes)
        if result.error:
            failed_files[file_name] = result.error
        elif result.row:
            successful_rows.append(result.row)
            expense_heads.update(result.expense_heads)
        else:
            failed_files[file_name] = "Unknown processing failure."

        progress.progress(index / len(files_to_process))

    consolidated_df = build_consolidated_dataframe(successful_rows, expense_heads)
    policy_df = build_policy_comparison_dataframe(successful_rows, consolidated_df)

    col1, col2, col3 = st.columns(3)
    col1.metric("Input Files", len(files_to_process) + len(precheck_errors))
    col2.metric("Successfully Processed", len(successful_rows))
    col3.metric("Failed / Skipped", len(failed_files))

    st.subheader("Consolidated Output")
    st.dataframe(consolidated_df, use_container_width=True)

    st.subheader("Policy Compliance")
    st.dataframe(policy_df, use_container_width=True)

    if failed_files:
        st.subheader("Failed / Skipped Files with Reason")
        st.json(failed_files)

    engine = _pick_excel_engine()
    if engine is None:
        st.warning(
            "Excel writer dependency missing (`xlsxwriter`/`openpyxl`). "
            "Providing CSV ZIP fallback download."
        )

        csv_zip = BytesIO()
        with ZipFile(csv_zip, "w") as zf:
            zf.writestr("consolidated.csv", consolidated_df.to_csv(index=False))
            zf.writestr("policy_check.csv", policy_df.to_csv(index=False))
        csv_zip.seek(0)

        st.download_button(
            "Download Report (CSV ZIP)",
            data=csv_zip,
            file_name="Dynamic_Expense_Report.zip",
            mime="application/zip",
        )
    else:
        output = BytesIO()
        with pd.ExcelWriter(output, engine=engine) as writer:
            consolidated_df.to_excel(writer, index=False, sheet_name="Consolidated")
            policy_df.to_excel(writer, index=False, sheet_name="Policy_Check")
            _style_output_workbook(writer, engine)
        output.seek(0)

        st.download_button(
            "Download Consolidated Report (.xlsx)",
            data=output,
            file_name="Dynamic_Expense_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
elif not uploaded_files:
    st.info("Upload one or more files and click 'Generate Consolidated Report'.")
